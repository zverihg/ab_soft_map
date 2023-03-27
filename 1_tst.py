#!/usr/bin/python
# -*- coding: utf-8 -*-
# encoding=utf8
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
from __future__ import unicode_literals
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
import os
import re
import sys
import openpyxl
from openpyxl import Workbook
import traceback
import	sqlite3 as sql
from datetime import datetime
from multiprocessing import Process
import requests
import lxml.html
from threading import Thread, Event
import time
from multiprocessing import Pool
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def dlog( tgt, msg ):		# w/o print

	with open ( tgt, 'a') as fil : fil.write( '{msg}\n'.format( msg = msg ) )

def inflo( msg ):	# with print
	tgt = './log_new_ref.txt'
	print (msg)
	with open ( tgt, 'a') as fil : fil.write( '{msg}\n'.format( msg = msg ) )

def dlog_prn( tgt, msg ):	# with print

	print (msg)
	with open ( tgt, 'a') as fil : fil.write( '{msg}\n'.format( msg = msg ) )

class tree:

	def __init__(self, itm):

		try:

			url = itm['lnk']
			nme = itm['nme']

			con = sql.connect( f'./{nme}.db' )
			cur = con.cursor()
			cur.execute('''drop table if exists tre_tbl; ''')
			cur.execute('''CREATE TABLE "tre_tbl" ("itm" TEXT, "parent" TEXT);''')
			cur.close()
			con.close()

			self.tag_loc = ['link', 'script', 'img', 'style', 'image']

			self.url_rot = url
			self.url_all = []
			self.tre	 = {}
			self.tre_don = {}
			self.sx = datetime.now()
			self.se = datetime.now()

			res = requests.get( url )
			dom_lst = str(res.content).split('\\n')
			for lnk in lxml.html.fromstring(str(res.content)).iterlinks():
				if 'http' not in lnk[-2] and lnk[0].tag not in ['link', 'script']: self.tre[lnk[-2]] = {}

			self.count = len(self.tre)
			print(self.count)

		except:	dlog_prn('__init___elo.txt', '{a}\n{b}\n\n'.format(a=traceback.format_tb(sys.exc_info()[2])[0],b=str(sys.exc_info()[1])))


	def set_aaa(self):
		self.url_all.append('dfsdfsdf')
		return self



	def set_tre(self, itm, idx):

		url = itm['lnk']
		nme = itm['nme']
		self.fff = 'fsdfsdfsfs'
		def rec_str(url, idx):

			lvl = 0

			def recursive(url, lvl):
				try:
					dct = {}
					# print(lvl)
					if lvl <= 1:
						res = requests.get( self.url_rot + url, timeout = 60 )
						time.sleep(0.5)
						dom_lst = str(res.content).split('\\n')
						hrf_lst = lxml.html.fromstring(str(res.content)).iterlinks()
						# print(self.url_rot + url)
						if not list(hrf_lst):
							cur.execute(f'''INSERT INTO tre_tbl (itm, parent) VALUES ('', '{url}');''')
							con.commit()
						else:
							for lnk in lxml.html.fromstring(str(res.content)).iterlinks():
								if 'http' not in lnk[-2] and 'www' not in lnk[-2] and lnk[-2] != '/' and lnk[0].tag not in self.tag_loc and lnk[-2] not in self.url_all:
									# print(lnk[0].tag)
									self.url_all.append(lnk[-2])
									# print(self.url_all)
									res = recursive(lnk[-2], lvl+1)
									dct[lnk[-2]] = res
									if res:
										for itm in set(res):
											cur.execute(f'''INSERT INTO tre_tbl (itm, parent) VALUES ('{itm}', '{lnk[-2]}');''')
											con.commit()
									else:
										cur.execute(f'''INSERT INTO tre_tbl (itm, parent) VALUES ('', '{lnk[-2]}');''')
										con.commit()
					return dct

				except:
					dlog_prn('recursive_elo.txt', '{a}\n{b}\n\n'.format(a=traceback.format_tb(sys.exc_info()[2])[0],b=str(sys.exc_info()[1])))
					dlog_prn('recursive_elo.txt', url)
					return {}

			con = sql.connect( f'./{nme}.db' )
			cur = con.cursor()

			res = recursive(url, lvl)
			for itm in res:
				cur.execute(f'''INSERT INTO tre_tbl (itm, parent) VALUES ('{itm}', '{url}');''')
				con.commit()
			# print(f'done {idx}')
			cur.close()
			con.close()

		idx = 0
		prc_lst = []
		for url, itm in self.tre.items(): prc_lst.append( Thread(target=rec_str, args=(url, idx,))) 

		for itm in prc_lst: itm.start()
		for itm in prc_lst: itm.join()

		self.se = datetime.now()
		return self


def main_old():

	os.system('clear')
	sx = datetime.now()

	url_lst = [ 
		# {'nme': 'crawler-test', 'lnk': 'http://crawler-test.com/'},
		# {'nme': 'stackoverflow', 'lnk': 'https://stackoverflow.com/'},
		# {'nme': 'google.com', 'lnk': 'https://google.com/'},
		# {'nme': 'dzen', 'lnk': 'https://dzen.ru/'},
		{'nme': 'dzen', 'lnk': 'https://zverihg.ru/tst'},
	]

	tre_lst = {}

	prc_lst = []

	for idx, itm in enumerate(url_lst):
		tre_lst[itm['nme']] = tree(itm = itm)
		# prc_lst.append( Process(target = tre_lst[itm['nme']].set_tre, args=(itm, idx,))) 
		prc_lst.append( Process( target = tre_lst[itm['nme']].set_aaa, args=() ) )

	for itm in prc_lst: itm.start()
	for itm in prc_lst: itm.join()

	se = datetime.now()

	print(se-sx)

	# print(tre_lst['dzen'].url)
	print(tre_lst['dzen'].se)
	print(tre_lst['dzen'].sx)
	print(tre_lst['dzen'].url_all)

	wbk = openpyxl.Workbook()
	sht = wbk.active
	hgh = sht.max_row

	sht['A1'].value = 'URL сайта'
	sht['B1'].value = 'Время обработки'
	sht['C1'].value = 'Кол-во найденных ссылок'
	sht['D1'].value = 'Имя файла с результатом'

	for idx, itm in enumerate(url_lst):
		# print(itm)
		tre_odj = tre_lst[itm['nme']]
		sht[f'A{idx + 2}'].value = itm["lnk"]
		sht[f'B{idx + 2}'].value = tre_odj.se 
		sht[f'C{idx + 2}'].value = len(tre_odj.url_all)
		sht[f'D{idx + 2}'].value = f'{itm["nme"]}.db'


	wbk.save('./stat.xlsx')

def main():

	os.system('clear')
	sx = datetime.now()

	url_lst = [ 
		{'nme': 'crawler-test', 'lnk': 'http://crawler-test.com/'},
		{'nme': 'stackoverflow', 'lnk': 'https://stackoverflow.com/'},
		{'nme': 'google.com', 'lnk': 'https://google.com/'},
		{'nme': 'dzen', 'lnk': 'https://dzen.ru/'},
		{'nme': 'vk', 'lnk': 'https://vk.com/'},
	]

	tre_lst = {}

	prc_lst = []
	pol_lst = []

	for idx, itm in enumerate(url_lst): pol_lst.append({'pol': Pool(processes=idx+1), 'itm': itm})

	for idx, itm in enumerate(pol_lst):
		tre_lst[itm['itm']['nme']] = tree(itm = itm['itm'])
		prc_lst.append({'prc': itm['pol'].apply_async(tre_lst[itm['itm']['nme']].set_tre, args=(itm['itm'], idx,)), 'nme': itm['itm']['nme']})

	for itm in prc_lst: tre_lst[itm['nme']] = itm['prc'].get()

	se = datetime.now()

	print(se-sx)

	print(tre_lst['dzen'].se)
	print(tre_lst['dzen'].sx)
	print(tre_lst['dzen'].url_all)

	wbk = openpyxl.Workbook()
	sht = wbk.active
	hgh = sht.max_row

	sht['A1'].value = 'URL сайта'
	sht['B1'].value = 'Время обработки'
	sht['C1'].value = 'Кол-во найденных ссылок'
	sht['D1'].value = 'Имя файла с результатом'

	for idx, itm in enumerate(url_lst):
		print(itm)
		tre_odj = tre_lst[itm['nme']]
		sht[f'A{idx + 2}'].value = itm["lnk"]
		sht[f'B{idx + 2}'].value = tre_odj.se - tre_odj.sx 
		sht[f'C{idx + 2}'].value = len(tre_odj.url_all)
		sht[f'D{idx + 2}'].value = f'{itm["nme"]}.db'


	wbk.save('./stat.xlsx')

if __name__ == "__main__":

	main()
